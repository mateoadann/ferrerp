"""Tests del servicio de importación masiva de productos."""

import csv
import io
from decimal import Decimal

from openpyxl import Workbook

from app.extensions import db
from app.models import Empresa
from app.models.categoria import Categoria
from app.models.movimiento_stock import MovimientoStock
from app.models.producto import Producto
from app.models.proveedor import Proveedor
from app.models.usuario import Usuario
from app.services.importacion_producto_service import (
    COLUMNAS_PLANTILLA,
    aplicar_importacion,
    generar_plantilla_excel,
    parsear_archivo,
    validar_importacion,
)

# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------


def _crear_empresa():
    """Crea y retorna una empresa de prueba."""
    empresa = Empresa(nombre='Ferretería Test', activa=True, aprobada=True)
    db.session.add(empresa)
    db.session.flush()
    return empresa


def _crear_usuario(empresa):
    """Crea y retorna un usuario administrador de prueba."""
    usuario = Usuario(
        email='admin@test.com',
        nombre='Admin Test',
        rol='administrador',
        activo=True,
        empresa_id=empresa.id,
    )
    usuario.set_password('clave123')
    db.session.add(usuario)
    db.session.flush()
    return usuario


def _crear_categoria(empresa, nombre, padre=None):
    """Crea y retorna una categoría de prueba."""
    cat = Categoria(
        nombre=nombre,
        empresa_id=empresa.id,
        padre_id=padre.id if padre else None,
        activa=True,
    )
    db.session.add(cat)
    db.session.flush()
    return cat


def _crear_proveedor(empresa, nombre='Proveedor Test'):
    """Crea y retorna un proveedor de prueba."""
    prov = Proveedor(
        nombre=nombre,
        empresa_id=empresa.id,
        activo=True,
    )
    db.session.add(prov)
    db.session.flush()
    return prov


def _crear_producto(empresa, codigo='PRD-001', nombre='Producto existente'):
    """Crea y retorna un producto de prueba."""
    prod = Producto(
        codigo=codigo,
        nombre=nombre,
        unidad_medida='unidad',
        precio_costo=Decimal('100.00'),
        precio_venta=Decimal('150.00'),
        stock_actual=Decimal('0'),
        stock_minimo=Decimal('5'),
        activo=True,
        empresa_id=empresa.id,
    )
    db.session.add(prod)
    db.session.flush()
    return prod


def _fila_valida(**overrides):
    """Retorna un dict con datos de fila válida para validación."""
    fila = {
        'codigo': 'FER-001',
        'nombre': 'Tornillo hexagonal',
        'descripcion': 'Tornillo 6x50',
        'categoria': 'Tornillería',
        'proveedor': 'Proveedor Test',
        'unidad_medida': 'unidad',
        'precio_costo': '150.00',
        'margen_ganancia': '40',
        'precio_venta': '210.00',
        'iva': '21',
        'stock_inicial': '100',
        'stock_minimo': '10',
        'codigo_barras': '7790001001001',
        'ubicacion': 'Pasillo 3',
    }
    fila.update(overrides)
    return fila


def _crear_xlsx_en_memoria(filas, headers=None):
    """Crea un archivo xlsx en memoria con las filas dadas. Retorna BytesIO."""
    wb = Workbook()
    ws = wb.active
    if headers is None:
        headers = COLUMNAS_PLANTILLA
    ws.append(headers)
    for fila in filas:
        row = [fila.get(h) for h in headers]
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _crear_csv_en_memoria(filas, headers=None):
    """Crea un archivo CSV en memoria con las filas dadas. Retorna BytesIO."""
    if headers is None:
        headers = COLUMNAS_PLANTILLA
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    for fila in filas:
        writer.writerow(fila)
    contenido = output.getvalue().encode('utf-8')
    return io.BytesIO(contenido)


class _FakeFileStorage:
    """Simula un FileStorage de Flask para tests del servicio."""

    def __init__(self, data: io.BytesIO):
        self._data = data

    def read(self):
        return self._data.read()


# --------------------------------------------------------------------------
# Tests unitarios: generar_plantilla_excel
# --------------------------------------------------------------------------


class TestGenerarPlantillaExcel:
    """Tests de generación de plantilla Excel."""

    def test_generar_plantilla_excel(self, app):
        """Genera una plantilla xlsx con headers, ejemplo y validaciones."""
        empresa = _crear_empresa()
        db.session.commit()

        resultado = generar_plantilla_excel(empresa.id)

        assert isinstance(resultado, io.BytesIO)

        from openpyxl import load_workbook

        wb = load_workbook(resultado)
        ws = wb.active

        # Verificar headers en fila 1
        headers_generados = [cell.value for cell in ws[1]]
        for col in COLUMNAS_PLANTILLA:
            assert col in headers_generados, f'Falta header: {col}'

        # Verificar fila de ejemplo (fila 2)
        fila_ejemplo = [cell.value for cell in ws[2]]
        assert fila_ejemplo[0] == 'FER-001'  # codigo
        assert fila_ejemplo[1] == 'Tornillo hexagonal 6x50'  # nombre

        # Verificar que existen data validations
        assert len(ws.data_validations.dataValidation) >= 2  # unidad_medida + iva mínimo

        wb.close()


# --------------------------------------------------------------------------
# Tests unitarios: parsear_archivo
# --------------------------------------------------------------------------


class TestParsearArchivo:
    """Tests de parseo de archivos xlsx y csv."""

    def test_parsear_archivo_xlsx(self, app):
        """Parsea un archivo xlsx y retorna lista de dicts correcta."""
        fila = _fila_valida()
        xlsx_data = _crear_xlsx_en_memoria([fila])
        storage = _FakeFileStorage(xlsx_data)

        resultado = parsear_archivo(storage, 'productos.xlsx')

        assert len(resultado) == 1
        assert resultado[0]['codigo'] == 'FER-001'
        assert resultado[0]['nombre'] == 'Tornillo hexagonal'
        assert resultado[0]['unidad_medida'] == 'unidad'

    def test_parsear_archivo_csv(self, app):
        """Parsea un archivo CSV y retorna lista de dicts correcta."""
        fila = _fila_valida()
        csv_data = _crear_csv_en_memoria([fila])
        storage = _FakeFileStorage(csv_data)

        resultado = parsear_archivo(storage, 'productos.csv')

        assert len(resultado) == 1
        assert resultado[0]['codigo'] == 'FER-001'
        assert resultado[0]['nombre'] == 'Tornillo hexagonal'

    def test_parsear_archivo_formato_invalido(self, app):
        """Rechaza archivos con extensión no soportada."""
        import pytest

        storage = _FakeFileStorage(io.BytesIO(b'contenido'))
        with pytest.raises(ValueError, match='Formato de archivo no soportado'):
            parsear_archivo(storage, 'datos.pdf')


# --------------------------------------------------------------------------
# Tests unitarios: validar_importacion — campos obligatorios
# --------------------------------------------------------------------------


class TestValidarCamposObligatorios:
    """Tests de validación de campos obligatorios."""

    def test_validar_campos_obligatorios(self, app):
        """Fila con campos obligatorios faltantes genera errores."""
        empresa = _crear_empresa()
        db.session.commit()

        fila_incompleta = {
            'codigo': 'FER-001',
            'nombre': None,
            'categoria': None,
            'unidad_medida': None,
            'precio_costo': None,
            'iva': None,
        }

        resultado = validar_importacion([fila_incompleta], empresa.id)

        assert len(resultado.errores) > 0
        campos_error = {e['campo'] for e in resultado.errores}
        assert 'nombre' in campos_error
        assert 'categoria' in campos_error
        assert 'unidad_medida' in campos_error
        assert 'precio_costo' in campos_error
        assert 'iva' in campos_error


# --------------------------------------------------------------------------
# Tests unitarios: validar_importacion — tipos inválidos
# --------------------------------------------------------------------------


class TestValidarTiposInvalidos:
    """Tests de validación de tipos numéricos."""

    def test_validar_tipos_invalidos(self, app):
        """precio_costo no numérico genera error."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila = _fila_valida(precio_costo='abc', precio_venta='xyz')

        resultado = validar_importacion([fila], empresa.id)

        assert len(resultado.errores) > 0
        campos_error = {e['campo'] for e in resultado.errores}
        assert 'precio_costo' in campos_error


# --------------------------------------------------------------------------
# Tests unitarios: validar_importacion — enums inválidos
# --------------------------------------------------------------------------


class TestValidarEnumsInvalidos:
    """Tests de validación de enums (unidad_medida, iva)."""

    def test_validar_enums_invalidos(self, app):
        """unidad_medida o iva inválidos generan error."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila = _fila_valida(unidad_medida='tonelada', iva='99')

        resultado = validar_importacion([fila], empresa.id)

        assert len(resultado.errores) > 0
        campos_error = {e['campo'] for e in resultado.errores}
        assert 'unidad_medida' in campos_error
        assert 'iva' in campos_error


# --------------------------------------------------------------------------
# Tests unitarios: validar_importacion — categorías
# --------------------------------------------------------------------------


class TestValidarCategorias:
    """Tests de validación de categorías existentes y nuevas."""

    def test_validar_categoria_existente(self, app):
        """Categoría que existe en la DB se matchea correctamente."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila = _fila_valida()

        resultado = validar_importacion([fila], empresa.id)

        assert len(resultado.filas_validas) == 1
        assert resultado.filas_validas[0]['categoria_id'] is not None

    def test_validar_categoria_no_existente_sin_crear(self, app):
        """Categoría inexistente genera advertencia cuando crear_categorias=False."""
        empresa = _crear_empresa()
        db.session.commit()

        fila = _fila_valida(categoria='Categoría Fantasma')

        resultado = validar_importacion([fila], empresa.id, crear_categorias=False)

        # La fila es válida pero sin categoría asignada
        assert len(resultado.filas_validas) == 1
        assert resultado.filas_validas[0]['categoria_id'] is None
        # Debe haber una advertencia
        assert len(resultado.advertencias) > 0
        mensajes = [a['mensaje'] for a in resultado.advertencias]
        assert any('Categoría Fantasma' in m for m in mensajes)

    def test_validar_categoria_no_existente_con_crear(self, app):
        """Categoría inexistente se detecta para auto-creación con crear_categorias=True."""
        empresa = _crear_empresa()
        db.session.commit()

        fila = _fila_valida(categoria='Categoría Nueva')

        resultado = validar_importacion([fila], empresa.id, crear_categorias=True)

        assert len(resultado.filas_validas) == 1
        assert 'Categoría Nueva' in resultado.categorias_nuevas


# --------------------------------------------------------------------------
# Tests unitarios: validar_importacion — duplicados
# --------------------------------------------------------------------------


class TestValidarDuplicados:
    """Tests de detección de duplicados en archivo y en la base de datos."""

    def test_validar_duplicado_en_archivo(self, app):
        """Dos filas con el mismo código en el archivo: la segunda es error."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila1 = _fila_valida(codigo='DUP-001')
        fila2 = _fila_valida(codigo='DUP-001', nombre='Duplicado')

        resultado = validar_importacion([fila1, fila2], empresa.id)

        assert len(resultado.filas_validas) == 1
        assert len(resultado.errores) >= 1
        mensajes_error = [e['mensaje'] for e in resultado.errores]
        assert any('duplicado' in m.lower() for m in mensajes_error)

    def test_validar_duplicado_en_db_saltar(self, app):
        """Producto existente con modo 'saltar' genera advertencia y se omite."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        _crear_producto(empresa, codigo='FER-001')
        db.session.commit()

        fila = _fila_valida(codigo='FER-001')

        resultado = validar_importacion([fila], empresa.id, modo_duplicados='saltar')

        assert len(resultado.filas_validas) == 0
        assert len(resultado.advertencias) >= 1
        mensajes = [a['mensaje'] for a in resultado.advertencias]
        assert any('ya existe' in m for m in mensajes)

    def test_validar_duplicado_en_db_actualizar(self, app):
        """Producto existente con modo 'actualizar' va a filas_actualizar."""
        empresa = _crear_empresa()
        _crear_categoria(empresa, 'Tornillería')
        producto = _crear_producto(empresa, codigo='FER-001')
        db.session.commit()

        fila = _fila_valida(codigo='FER-001')

        resultado = validar_importacion([fila], empresa.id, modo_duplicados='actualizar')

        assert len(resultado.filas_actualizar) == 1
        assert resultado.filas_actualizar[0]['producto_existente_id'] == producto.id


# --------------------------------------------------------------------------
# Tests unitarios: aplicar_importacion
# --------------------------------------------------------------------------


class TestAplicarImportacion:
    """Tests de aplicación de importación (creación de productos en DB)."""

    def test_aplicar_importacion_crea_productos(self, app):
        """aplicar_importacion crea los productos en la base de datos."""
        empresa = _crear_empresa()
        usuario = _crear_usuario(empresa)
        cat = _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila = _fila_valida()
        resultado = validar_importacion([fila], empresa.id)
        assert len(resultado.filas_validas) == 1

        auditoria = aplicar_importacion(
            resultado,
            empresa.id,
            usuario.id,
            nombre_archivo='test.xlsx',
        )
        db.session.commit()

        # Verificar producto creado
        producto = Producto.query.filter_by(empresa_id=empresa.id, codigo='FER-001').first()
        assert producto is not None
        assert producto.nombre == 'Tornillo hexagonal'
        assert producto.categoria_id == cat.id
        assert producto.precio_costo == Decimal('150.00')

        # Verificar auditoría
        assert auditoria.filas_importadas == 1
        assert auditoria.nombre_archivo == 'test.xlsx'

    def test_aplicar_importacion_crea_categorias(self, app):
        """aplicar_importacion con crear_categorias=True crea categorías nuevas."""
        empresa = _crear_empresa()
        usuario = _crear_usuario(empresa)
        db.session.commit()

        fila = _fila_valida(categoria='Herramientas Nuevas')
        resultado = validar_importacion([fila], empresa.id, crear_categorias=True)
        assert 'Herramientas Nuevas' in resultado.categorias_nuevas

        auditoria = aplicar_importacion(
            resultado,
            empresa.id,
            usuario.id,
            crear_categorias=True,
            nombre_archivo='test.xlsx',
        )
        db.session.commit()

        # Verificar categoría creada
        cat = Categoria.query.filter_by(empresa_id=empresa.id, nombre='Herramientas Nuevas').first()
        assert cat is not None

        # Verificar producto asociado a la categoría
        producto = Producto.query.filter_by(empresa_id=empresa.id, codigo='FER-001').first()
        assert producto is not None
        assert producto.categoria_id == cat.id

        # Verificar auditoría
        assert auditoria.categorias_creadas >= 1

    def test_aplicar_importacion_stock_inicial(self, app):
        """aplicar_importacion con stock_inicial > 0 crea MovimientoStock."""
        empresa = _crear_empresa()
        usuario = _crear_usuario(empresa)
        _crear_categoria(empresa, 'Tornillería')
        db.session.commit()

        fila = _fila_valida(stock_inicial='50')
        resultado = validar_importacion([fila], empresa.id)
        assert len(resultado.filas_validas) == 1

        aplicar_importacion(
            resultado,
            empresa.id,
            usuario.id,
            nombre_archivo='test.xlsx',
        )
        db.session.commit()

        # Verificar producto con stock
        producto = Producto.query.filter_by(empresa_id=empresa.id, codigo='FER-001').first()
        assert producto is not None
        assert producto.stock_actual == Decimal('50')

        # Verificar movimiento de stock
        movimiento = MovimientoStock.query.filter_by(producto_id=producto.id).first()
        assert movimiento is not None
        assert movimiento.tipo == 'ajuste_positivo'
        assert movimiento.cantidad == Decimal('50')
        assert movimiento.referencia_tipo == 'importacion'
