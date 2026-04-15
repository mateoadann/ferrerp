"""Servicio de importación masiva de productos."""

import csv
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..extensions import db
from ..models.categoria import Categoria
from ..models.importacion_producto import ImportacionProducto
from ..models.movimiento_stock import MovimientoStock
from ..models.producto import Producto
from ..models.proveedor import Proveedor

MAX_FILAS = 5000

COLUMNAS_PLANTILLA = [
    'codigo',
    'nombre',
    'descripcion',
    'categoria',
    'proveedor',
    'unidad_medida',
    'precio_costo',
    'margen_ganancia',
    'precio_venta',
    'iva',
    'stock_inicial',
    'stock_minimo',
    'codigo_barras',
    'ubicacion',
]

ANCHOS_COLUMNA = {
    'codigo': 14,
    'nombre': 30,
    'descripcion': 35,
    'categoria': 25,
    'proveedor': 25,
    'unidad_medida': 16,
    'precio_costo': 14,
    'margen_ganancia': 16,
    'precio_venta': 14,
    'iva': 10,
    'stock_inicial': 14,
    'stock_minimo': 14,
    'codigo_barras': 18,
    'ubicacion': 16,
}


@dataclass
class ImportacionResult:
    """Resultado de la validación de una importación de productos."""

    filas_validas: list[dict] = field(default_factory=list)
    filas_actualizar: list[dict] = field(default_factory=list)
    errores: list[dict] = field(default_factory=list)
    advertencias: list[dict] = field(default_factory=list)
    categorias_nuevas: list[str] = field(default_factory=list)


def _obtener_categorias_empresa(empresa_id: int) -> list[str]:
    """Obtiene los nombres completos de las categorías activas de la empresa."""
    categorias = Categoria.query.filter_by(empresa_id=empresa_id, activa=True).all()
    return sorted([c.nombre_completo for c in categorias])


def _obtener_proveedores_empresa(empresa_id: int) -> list[str]:
    """Obtiene los nombres de los proveedores activos de la empresa."""
    proveedores = (
        Proveedor.query.filter_by(empresa_id=empresa_id, activo=True)
        .order_by(Proveedor.nombre)
        .all()
    )
    return [p.nombre for p in proveedores]


def generar_plantilla_excel(empresa_id: int) -> BytesIO:
    """Genera una plantilla Excel (.xlsx) con headers, validaciones y ejemplo.

    Args:
        empresa_id: ID de la empresa para cargar categorías dinámicas.

    Returns:
        BytesIO con el contenido del archivo xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # --- Estilos de encabezado ---
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Escribir encabezados ---
    for col_idx, col_name in enumerate(COLUMNAS_PLANTILLA, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Anchos de columna ---
    for col_idx, col_name in enumerate(COLUMNAS_PLANTILLA, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = ANCHOS_COLUMNA.get(col_name, 14)

    # --- Fila de ejemplo ---
    ejemplo = [
        'FER-001',  # codigo
        'Tornillo hexagonal 6x50',  # nombre
        'Tornillo cabeza hexagonal zincado 6mm x 50mm',  # descripcion
        'Tornillería',  # categoria
        'Bulonera del Sur',  # proveedor
        'unidad',  # unidad_medida
        150.00,  # precio_costo
        40,  # margen_ganancia (%)
        210.00,  # precio_venta
        21,  # iva
        100,  # stock_inicial
        10,  # stock_minimo
        '7790001001001',  # codigo_barras
        'Pasillo 3 - Estante B',  # ubicacion
    ]
    ws.append(ejemplo)

    # --- Data Validations ---
    # Rango de celdas para validación (filas 2 a 5001 para cubrir max filas)
    rango_validacion = MAX_FILAS + 1

    # Unidad de medida
    col_unidad = get_column_letter(COLUMNAS_PLANTILLA.index('unidad_medida') + 1)
    dv_unidad = DataValidation(
        type='list',
        formula1='"unidad,metro,kilo,litro,par"',
        allow_blank=True,
    )
    dv_unidad.error = 'Valor inválido. Use: unidad, metro, kilo, litro, par'
    dv_unidad.errorTitle = 'Unidad de medida inválida'
    dv_unidad.prompt = 'Seleccione la unidad de medida'
    dv_unidad.promptTitle = 'Unidad de medida'
    ws.add_data_validation(dv_unidad)
    dv_unidad.add(f'{col_unidad}2:{col_unidad}{rango_validacion}')

    # IVA
    col_iva = get_column_letter(COLUMNAS_PLANTILLA.index('iva') + 1)
    dv_iva = DataValidation(
        type='list',
        formula1='"21,10.5,27"',
        allow_blank=True,
    )
    dv_iva.error = 'Valor inválido. Use: 21, 10.5, 27'
    dv_iva.errorTitle = 'IVA inválido'
    dv_iva.prompt = 'Seleccione el porcentaje de IVA'
    dv_iva.promptTitle = 'IVA'
    ws.add_data_validation(dv_iva)
    dv_iva.add(f'{col_iva}2:{col_iva}{rango_validacion}')

    # Categorías dinámicas desde la DB
    nombres_categorias = _obtener_categorias_empresa(empresa_id)
    if nombres_categorias:
        # DataValidation con lista tiene límite de ~255 chars en formula1
        lista_cats = ','.join(nombres_categorias)
        if len(lista_cats) <= 255:
            col_cat = get_column_letter(COLUMNAS_PLANTILLA.index('categoria') + 1)
            dv_cat = DataValidation(
                type='list',
                formula1=f'"{lista_cats}"',
                allow_blank=True,
            )
            dv_cat.error = 'Categoría no encontrada. Puede escribir el nombre manualmente.'
            dv_cat.errorTitle = 'Categoría'
            dv_cat.showErrorMessage = False  # Permitir valores fuera de la lista
            dv_cat.prompt = 'Seleccione o escriba la categoría (formato: Padre > Hijo)'
            dv_cat.promptTitle = 'Categoría'
            ws.add_data_validation(dv_cat)
            dv_cat.add(f'{col_cat}2:{col_cat}{rango_validacion}')

    # Proveedores dinámicos desde la DB
    nombres_proveedores = _obtener_proveedores_empresa(empresa_id)
    if nombres_proveedores:
        lista_provs = ','.join(nombres_proveedores)
        if len(lista_provs) <= 255:
            col_prov = get_column_letter(COLUMNAS_PLANTILLA.index('proveedor') + 1)
            dv_prov = DataValidation(
                type='list',
                formula1=f'"{lista_provs}"',
                allow_blank=True,
            )
            dv_prov.error = 'Proveedor no encontrado.'
            dv_prov.errorTitle = 'Proveedor'
            dv_prov.showErrorMessage = False
            dv_prov.prompt = 'Seleccione el proveedor'
            dv_prov.promptTitle = 'Proveedor'
            ws.add_data_validation(dv_prov)
            dv_prov.add(f'{col_prov}2:{col_prov}{rango_validacion}')

    # --- Congelar fila de encabezado ---
    ws.freeze_panes = 'A2'

    # --- Guardar a BytesIO ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _normalizar_nombre_columna(nombre: str) -> str:
    """Normaliza un nombre de columna: minúsculas, sin espacios extra, _ en vez de espacio."""
    if nombre is None:
        return ''
    return str(nombre).strip().lower().replace(' ', '_')


def _valor_celda(valor):
    """Convierte un valor de celda a su representación Python adecuada."""
    if valor is None:
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        return valor if valor else None
    return valor


def _convertir_decimal(valor, nombre_campo: str) -> Decimal | None:
    """Convierte un valor a Decimal de forma segura. Retorna None si está vacío."""
    if valor is None:
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        raise ValueError(f'El campo {nombre_campo} debe ser numérico: "{valor}"')


def _fila_esta_vacia(fila: dict) -> bool:
    """Verifica si una fila tiene todos sus valores vacíos/None."""
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in fila.values())


def _parsear_xlsx(file_storage) -> list[dict]:
    """Parsea un archivo Excel y retorna lista de dicts."""
    contenido = BytesIO(file_storage.read())
    wb = load_workbook(contenido, read_only=True, data_only=True)
    ws = wb.active

    filas = []
    headers = None

    for idx_fila, fila in enumerate(ws.iter_rows(values_only=True)):
        if idx_fila == 0:
            headers = [_normalizar_nombre_columna(h) for h in fila]
            continue

        if len(filas) >= MAX_FILAS:
            raise ValueError(
                f'El archivo excede el límite de {MAX_FILAS} filas. '
                'Divida el archivo en partes más pequeñas.'
            )

        fila_dict = {}
        for col_idx, valor in enumerate(fila):
            if col_idx < len(headers) and headers[col_idx]:
                fila_dict[headers[col_idx]] = _valor_celda(valor)

        if not _fila_esta_vacia(fila_dict):
            filas.append(fila_dict)

    wb.close()
    return filas


def _parsear_csv(file_storage) -> list[dict]:
    """Parsea un archivo CSV, intentando UTF-8 primero y Latin-1 como fallback."""
    contenido_bytes = file_storage.read()

    # Intentar UTF-8 primero, luego Latin-1
    contenido_texto = None
    for encoding in ('utf-8', 'latin-1'):
        try:
            contenido_texto = contenido_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue

    if contenido_texto is None:
        raise ValueError(
            'No se pudo decodificar el archivo CSV. '
            'Asegúrese de que esté en formato UTF-8 o Latin-1.'
        )

    # Detectar delimitador (coma o punto y coma)
    primera_linea = contenido_texto.split('\n')[0] if contenido_texto else ''
    delimitador = ';' if primera_linea.count(';') > primera_linea.count(',') else ','

    reader = csv.DictReader(io.StringIO(contenido_texto), delimiter=delimitador)

    # Normalizar nombres de columna del header
    if reader.fieldnames:
        reader.fieldnames = [_normalizar_nombre_columna(f) for f in reader.fieldnames]

    filas = []
    for fila in reader:
        if len(filas) >= MAX_FILAS:
            raise ValueError(
                f'El archivo excede el límite de {MAX_FILAS} filas. '
                'Divida el archivo en partes más pequeñas.'
            )

        fila_limpia = {k: _valor_celda(v) for k, v in fila.items() if k}

        if not _fila_esta_vacia(fila_limpia):
            filas.append(fila_limpia)

    return filas


def parsear_archivo(file_storage, nombre_archivo: str) -> list[dict]:
    """Parsea un archivo Excel o CSV y retorna las filas como lista de dicts.

    Args:
        file_storage: objeto FileStorage de Flask.
        nombre_archivo: nombre original del archivo para detectar tipo.

    Returns:
        Lista de diccionarios, uno por fila del archivo.

    Raises:
        ValueError: si el archivo excede el límite de filas o tiene formato inválido.
    """
    if not nombre_archivo:
        raise ValueError('Nombre de archivo no proporcionado.')

    extension = nombre_archivo.rsplit('.', 1)[-1].lower() if '.' in nombre_archivo else ''

    if extension == 'xlsx':
        return _parsear_xlsx(file_storage)
    elif extension == 'csv':
        return _parsear_csv(file_storage)
    else:
        raise ValueError(
            f'Formato de archivo no soportado: .{extension}. ' 'Use archivos .xlsx o .csv.'
        )


CAMPOS_OBLIGATORIOS = ['codigo', 'nombre', 'categoria', 'unidad_medida', 'precio_costo', 'iva']
UNIDADES_VALIDAS = {'unidad', 'metro', 'kilo', 'litro', 'par'}
IVA_VALIDOS = {Decimal('21'), Decimal('10.5'), Decimal('27')}
CAMPOS_NUMERICOS = [
    'precio_costo',
    'margen_ganancia',
    'precio_venta',
    'stock_inicial',
    'stock_minimo',
]


def _cargar_categorias_mapa(empresa_id: int) -> dict[str, Categoria]:
    """Carga categorías activas y retorna mapa {nombre_completo_lower: Categoria}."""
    categorias = Categoria.query.filter_by(empresa_id=empresa_id, activa=True).all()
    return {c.nombre_completo.lower(): c for c in categorias}


def _cargar_proveedores_mapa(empresa_id: int) -> dict[str, Proveedor]:
    """Carga proveedores activos y retorna mapa {nombre_lower: Proveedor}."""
    proveedores = Proveedor.query.filter_by(empresa_id=empresa_id, activo=True).all()
    return {p.nombre.lower(): p for p in proveedores}


def _cargar_productos_existentes(empresa_id: int) -> dict[str, Producto]:
    """Carga productos existentes y retorna mapa {codigo_lower: Producto}."""
    productos = Producto.query.filter_by(empresa_id=empresa_id).all()
    return {p.codigo.lower(): p for p in productos}


def _validar_campos_obligatorios(fila: dict, num_fila: int, errores: list) -> bool:
    """Valida que estén presentes los campos obligatorios. Retorna True si ok."""
    tiene_error = False
    for campo in CAMPOS_OBLIGATORIOS:
        valor = fila.get(campo)
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            errores.append(
                {
                    'fila': num_fila,
                    'campo': campo,
                    'mensaje': f'{campo} es obligatorio',
                }
            )
            tiene_error = True
    return not tiene_error


def _validar_tipos_numericos(fila: dict, num_fila: int, errores: list) -> dict[str, Decimal]:
    """Valida y convierte campos numéricos. Retorna dict de valores convertidos."""
    valores = {}
    for campo in CAMPOS_NUMERICOS:
        valor_raw = fila.get(campo)
        if valor_raw is None or (isinstance(valor_raw, str) and not valor_raw.strip()):
            valores[campo] = None
            continue
        try:
            valor_dec = Decimal(str(valor_raw))
        except (InvalidOperation, ValueError):
            errores.append(
                {
                    'fila': num_fila,
                    'campo': campo,
                    'mensaje': f'{campo} debe ser un número válido: "{valor_raw}"',
                }
            )
            valores[campo] = None
            continue

        if valor_dec < 0:
            errores.append(
                {
                    'fila': num_fila,
                    'campo': campo,
                    'mensaje': f'{campo} no puede ser negativo',
                }
            )
            valores[campo] = None
            continue

        valores[campo] = valor_dec
    return valores


def _validar_enums(fila: dict, num_fila: int, errores: list) -> bool:
    """Valida unidad_medida e IVA. Retorna True si ok."""
    tiene_error = False
    unidad = str(fila.get('unidad_medida', '')).strip().lower()
    if unidad and unidad not in UNIDADES_VALIDAS:
        errores.append(
            {
                'fila': num_fila,
                'campo': 'unidad_medida',
                'mensaje': (
                    f'Unidad de medida inválida: "{unidad}". '
                    f'Valores permitidos: {", ".join(sorted(UNIDADES_VALIDAS))}'
                ),
            }
        )
        tiene_error = True

    iva_raw = fila.get('iva')
    if iva_raw is not None:
        try:
            iva_val = Decimal(str(iva_raw))
            if iva_val not in IVA_VALIDOS:
                errores.append(
                    {
                        'fila': num_fila,
                        'campo': 'iva',
                        'mensaje': f'IVA inválido: {iva_raw}. Valores permitidos: 21, 10.5, 27',
                    }
                )
                tiene_error = True
        except (InvalidOperation, ValueError):
            errores.append(
                {
                    'fila': num_fila,
                    'campo': 'iva',
                    'mensaje': f'IVA debe ser numérico: "{iva_raw}"',
                }
            )
            tiene_error = True

    return not tiene_error


def _calcular_precio_venta(
    precio_costo: Decimal | None,
    margen: Decimal | None,
    precio_venta_raw: Decimal | None,
) -> Decimal | None:
    """Calcula precio_venta si no fue provisto. No incluye IVA (se guarda neto)."""
    if precio_venta_raw is not None:
        return precio_venta_raw
    if precio_costo is not None and margen is not None:
        return precio_costo * (1 + margen / Decimal('100'))
    return precio_costo


def validar_importacion(
    filas: list[dict],
    empresa_id: int,
    crear_categorias: bool = False,
    modo_duplicados: str = 'saltar',
) -> ImportacionResult:
    """Valida las filas parseadas contra reglas de negocio y datos existentes.

    Args:
        filas: lista de dicts con los datos de cada fila.
        empresa_id: ID de la empresa para validar contra productos existentes.
        crear_categorias: si True, detecta categorías nuevas para auto-creación.
        modo_duplicados: 'saltar', 'actualizar' o 'error'.

    Returns:
        ImportacionResult con filas clasificadas y errores/advertencias.
    """
    resultado = ImportacionResult()

    # Pre-cargar datos de referencia
    categorias_mapa = _cargar_categorias_mapa(empresa_id)
    proveedores_mapa = _cargar_proveedores_mapa(empresa_id)
    productos_existentes = _cargar_productos_existentes(empresa_id)

    # Tracking de códigos vistos en el archivo para detectar duplicados internos
    codigos_vistos: dict[str, int] = {}
    categorias_nuevas_set: set[str] = set()

    for idx, fila in enumerate(filas):
        num_fila = idx + 2  # +2 porque fila 1 es el header en Excel
        errores_fila: list[dict] = []

        # 1. Campos obligatorios
        _validar_campos_obligatorios(fila, num_fila, errores_fila)

        # 2. Validación de tipos numéricos
        valores_num = _validar_tipos_numericos(fila, num_fila, errores_fila)

        # 3. Validación de enums
        _validar_enums(fila, num_fila, errores_fila)

        # Si hay errores de validación básica, registrar y continuar
        if errores_fila:
            resultado.errores.extend(errores_fila)
            continue

        codigo = str(fila.get('codigo', '')).strip()
        codigo_lower = codigo.lower()

        # 5. Duplicado dentro del archivo
        if codigo_lower in codigos_vistos:
            resultado.errores.append(
                {
                    'fila': num_fila,
                    'campo': 'codigo',
                    'mensaje': (
                        f'Código duplicado en archivo: "{codigo}" '
                        f'(primera aparición en fila {codigos_vistos[codigo_lower]})'
                    ),
                }
            )
            continue
        codigos_vistos[codigo_lower] = num_fila

        # 6. Duplicado contra DB
        producto_existente = productos_existentes.get(codigo_lower)
        if producto_existente:
            if modo_duplicados == 'saltar':
                resultado.advertencias.append(
                    {
                        'fila': num_fila,
                        'mensaje': (
                            f'Producto con código "{codigo}" ya existe, se omite (modo: saltar)'
                        ),
                    }
                )
                continue
            elif modo_duplicados == 'error':
                resultado.errores.append(
                    {
                        'fila': num_fila,
                        'campo': 'codigo',
                        'mensaje': f'Producto con código "{codigo}" ya existe en la base de datos',
                    }
                )
                continue
            # modo_duplicados == 'actualizar' → continuar para agregar a filas_actualizar

        # 3b. Matching de categoría
        categoria_nombre = str(fila.get('categoria', '')).strip()
        categoria_obj = categorias_mapa.get(categoria_nombre.lower())
        categoria_id = categoria_obj.id if categoria_obj else None

        if not categoria_obj:
            if crear_categorias:
                categorias_nuevas_set.add(categoria_nombre)
            else:
                resultado.advertencias.append(
                    {
                        'fila': num_fila,
                        'mensaje': (
                            f'Categoría "{categoria_nombre}" no encontrada. '
                            'El producto se importará sin categoría.'
                        ),
                    }
                )

        # 4. Matching de proveedor
        proveedor_nombre = fila.get('proveedor')
        proveedor_id = None
        if proveedor_nombre:
            proveedor_nombre = str(proveedor_nombre).strip()
            proveedor_obj = proveedores_mapa.get(proveedor_nombre.lower())
            if proveedor_obj:
                proveedor_id = proveedor_obj.id
            else:
                resultado.advertencias.append(
                    {
                        'fila': num_fila,
                        'mensaje': (
                            f'Proveedor "{proveedor_nombre}" no encontrado. '
                            'El producto se importará sin proveedor.'
                        ),
                    }
                )

        # 7. Calcular precio_venta si no provisto
        precio_costo = valores_num.get('precio_costo')
        margen = valores_num.get('margen_ganancia')
        precio_venta = _calcular_precio_venta(precio_costo, margen, valores_num.get('precio_venta'))

        # Construir fila validada
        fila_validada = {
            'fila_origen': num_fila,
            'codigo': codigo,
            'nombre': str(fila.get('nombre', '')).strip(),
            'descripcion': str(fila.get('descripcion', '')).strip()
            if fila.get('descripcion')
            else None,
            'categoria_id': categoria_id,
            'categoria_nombre': categoria_nombre,
            'proveedor_id': proveedor_id,
            'unidad_medida': str(fila.get('unidad_medida', 'unidad')).strip().lower(),
            'precio_costo': precio_costo or Decimal('0'),
            'precio_venta': precio_venta or Decimal('0'),
            'iva_porcentaje': Decimal(str(fila.get('iva', '21'))),
            'stock_inicial': valores_num.get('stock_inicial') or Decimal('0'),
            'stock_minimo': valores_num.get('stock_minimo') or Decimal('0'),
            'codigo_barras': str(fila.get('codigo_barras', '')).strip()
            if fila.get('codigo_barras')
            else None,
            'ubicacion': str(fila.get('ubicacion', '')).strip() if fila.get('ubicacion') else None,
        }

        if producto_existente and modo_duplicados == 'actualizar':
            fila_validada['producto_existente_id'] = producto_existente.id
            resultado.filas_actualizar.append(fila_validada)
        else:
            resultado.filas_validas.append(fila_validada)

    resultado.categorias_nuevas = sorted(categorias_nuevas_set)
    return resultado


def _crear_categorias_nuevas(nombres: list[str], empresa_id: int) -> dict[str, int]:
    """Crea categorías nuevas respetando jerarquía 'Padre > Hijo'.

    Retorna mapa {nombre_completo_lower: categoria_id} de las categorías creadas.
    """
    creadas: dict[str, int] = {}

    # Primero separar padres y subcategorías
    padres_necesarios: set[str] = set()
    subcategorias: list[tuple[str, str]] = []  # (padre, hijo)

    for nombre in nombres:
        if ' > ' in nombre:
            partes = nombre.split(' > ', 1)
            padre_nombre = partes[0].strip()
            hijo_nombre = partes[1].strip()
            padres_necesarios.add(padre_nombre)
            subcategorias.append((padre_nombre, hijo_nombre))
        else:
            padres_necesarios.add(nombre)

    # Cargar categorías existentes para no duplicar
    cats_existentes = Categoria.query.filter_by(empresa_id=empresa_id, activa=True).all()
    mapa_padres: dict[str, Categoria] = {}
    for c in cats_existentes:
        if c.padre_id is None:
            mapa_padres[c.nombre.lower()] = c

    # Crear padres que faltan
    for padre_nombre in padres_necesarios:
        if padre_nombre.lower() not in mapa_padres:
            nueva_cat = Categoria(
                nombre=padre_nombre,
                empresa_id=empresa_id,
                activa=True,
            )
            db.session.add(nueva_cat)
            db.session.flush()
            mapa_padres[padre_nombre.lower()] = nueva_cat
            creadas[padre_nombre.lower()] = nueva_cat.id

    # Crear subcategorías
    for padre_nombre, hijo_nombre in subcategorias:
        padre_cat = mapa_padres.get(padre_nombre.lower())
        if not padre_cat:
            continue

        # Verificar si la subcategoría ya existe
        nombre_completo = f'{padre_nombre} > {hijo_nombre}'
        sub_existente = Categoria.query.filter_by(
            empresa_id=empresa_id,
            nombre=hijo_nombre,
            padre_id=padre_cat.id,
        ).first()

        if sub_existente:
            creadas[nombre_completo.lower()] = sub_existente.id
        else:
            nueva_sub = Categoria(
                nombre=hijo_nombre,
                padre_id=padre_cat.id,
                empresa_id=empresa_id,
                activa=True,
            )
            db.session.add(nueva_sub)
            db.session.flush()
            creadas[nombre_completo.lower()] = nueva_sub.id

    # También registrar padres que no son subcategoría
    for padre_nombre in padres_necesarios:
        nombre_lower = padre_nombre.lower()
        if nombre_lower not in creadas:
            padre_cat = mapa_padres.get(nombre_lower)
            if padre_cat:
                creadas[nombre_lower] = padre_cat.id

    return creadas


def _crear_movimiento_stock(
    producto: Producto,
    cantidad: Decimal,
    empresa_id: int,
    usuario_id: int,
) -> MovimientoStock:
    """Crea un MovimientoStock de tipo ajuste_positivo para stock inicial."""
    stock_anterior_val, stock_posterior_val = producto.actualizar_stock(cantidad, 'ajuste_positivo')
    movimiento = MovimientoStock(
        producto_id=producto.id,
        tipo='ajuste_positivo',
        cantidad=cantidad,
        stock_anterior=stock_anterior_val,
        stock_posterior=stock_posterior_val,
        referencia_tipo='importacion',
        motivo='Importación masiva de productos',
        usuario_id=usuario_id,
        empresa_id=empresa_id,
    )
    db.session.add(movimiento)
    return movimiento


def aplicar_importacion(
    resultado: ImportacionResult,
    empresa_id: int,
    usuario_id: int,
    crear_categorias: bool = False,
    nombre_archivo: str = '',
) -> ImportacionProducto:
    """Aplica la importación validada creando productos en una transacción.

    No hace commit — el caller es responsable de hacer db.session.commit().

    Args:
        resultado: ImportacionResult de la validación previa.
        empresa_id: ID de la empresa.
        usuario_id: ID del usuario que ejecuta la importación.
        crear_categorias: si True, crea categorías inexistentes.
        nombre_archivo: nombre del archivo original para auditoría.

    Returns:
        Instancia de ImportacionProducto (registro de auditoría).
    """
    categorias_creadas_count = 0
    categorias_nuevas_mapa: dict[str, int] = {}

    # Crear categorías nuevas si se solicitó
    if crear_categorias and resultado.categorias_nuevas:
        categorias_nuevas_mapa = _crear_categorias_nuevas(resultado.categorias_nuevas, empresa_id)
        categorias_creadas_count = len(categorias_nuevas_mapa)

    # Crear productos nuevos
    for fila in resultado.filas_validas:
        categoria_id = fila['categoria_id']
        # Si la categoría era nueva y se creó, asignar el id
        if categoria_id is None and fila.get('categoria_nombre'):
            cat_nombre_lower = fila['categoria_nombre'].lower()
            categoria_id = categorias_nuevas_mapa.get(cat_nombre_lower)

        producto = Producto(
            codigo=fila['codigo'],
            nombre=fila['nombre'],
            descripcion=fila.get('descripcion'),
            categoria_id=categoria_id,
            proveedor_id=fila.get('proveedor_id'),
            unidad_medida=fila['unidad_medida'],
            precio_costo=fila['precio_costo'],
            precio_venta=fila['precio_venta'],
            iva_porcentaje=fila['iva_porcentaje'],
            stock_actual=Decimal('0'),
            stock_minimo=fila.get('stock_minimo', Decimal('0')),
            codigo_barras=fila.get('codigo_barras'),
            ubicacion=fila.get('ubicacion'),
            activo=True,
            empresa_id=empresa_id,
        )
        db.session.add(producto)
        db.session.flush()  # Para obtener producto.id

        # Stock inicial
        stock_inicial = fila.get('stock_inicial', Decimal('0'))
        if stock_inicial and stock_inicial > 0:
            _crear_movimiento_stock(producto, stock_inicial, empresa_id, usuario_id)

    # Actualizar productos existentes
    for fila in resultado.filas_actualizar:
        producto = db.session.get(Producto, fila['producto_existente_id'])
        if not producto:
            continue

        # Actualizar solo campos con valor
        if fila.get('nombre'):
            producto.nombre = fila['nombre']
        if fila.get('descripcion') is not None:
            producto.descripcion = fila['descripcion']

        categoria_id = fila.get('categoria_id')
        if categoria_id is None and fila.get('categoria_nombre'):
            cat_nombre_lower = fila['categoria_nombre'].lower()
            categoria_id = categorias_nuevas_mapa.get(cat_nombre_lower)
        if categoria_id is not None:
            producto.categoria_id = categoria_id

        if fila.get('proveedor_id') is not None:
            producto.proveedor_id = fila['proveedor_id']
        if fila.get('unidad_medida'):
            producto.unidad_medida = fila['unidad_medida']
        if fila.get('precio_costo') and fila['precio_costo'] > 0:
            producto.precio_costo = fila['precio_costo']
        if fila.get('precio_venta') and fila['precio_venta'] > 0:
            producto.precio_venta = fila['precio_venta']
        if fila.get('iva_porcentaje'):
            producto.iva_porcentaje = fila['iva_porcentaje']
        if fila.get('stock_minimo') and fila['stock_minimo'] > 0:
            producto.stock_minimo = fila['stock_minimo']
        if fila.get('codigo_barras'):
            producto.codigo_barras = fila['codigo_barras']
        if fila.get('ubicacion'):
            producto.ubicacion = fila['ubicacion']

        # Stock inicial como ajuste adicional
        stock_inicial = fila.get('stock_inicial', Decimal('0'))
        if stock_inicial and stock_inicial > 0:
            _crear_movimiento_stock(producto, stock_inicial, empresa_id, usuario_id)

    # Crear registro de auditoría
    total_filas = (
        len(resultado.filas_validas) + len(resultado.filas_actualizar) + len(resultado.errores)
    )
    auditoria = ImportacionProducto(
        usuario_id=usuario_id,
        empresa_id=empresa_id,
        nombre_archivo=nombre_archivo,
        total_filas=total_filas,
        filas_importadas=len(resultado.filas_validas),
        filas_actualizadas=len(resultado.filas_actualizar),
        filas_omitidas=len(resultado.advertencias),
        errores_count=len(resultado.errores),
        categorias_creadas=categorias_creadas_count,
        modo_duplicados='saltar',  # Se podría pasar como parámetro
    )
    db.session.add(auditoria)

    return auditoria
